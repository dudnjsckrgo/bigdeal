# -*- coding: utf-8 -*-
"""
Created on Thu Aug 29 14:26:13 2019

@author: Administrator
"""
from selenium import webdriver
from time import sleep
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
import pyautogui
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.by import By


#exel_path ='/Users/Administrator/Downloads/온라인판매20190726.xlsx'
import pyperclip
from excel import Excel
#data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
#load_wb = load_workbook(exel_path , data_only=True)
#시트 이름으로 불러오기
#load_ws = load_wb['Sheet1']
 
#셀 주소로 값 출력
#print(load_ws['B5'].value)
 
#for i in range(6,189):
#i=6
#str_1=load_ws.cell(i,2).value
#str_2=load_ws.cell(i,1).value


class Itemscout:
    
    
    def __init__(self):
        self.item_name=''
        self.category =''
        self.driver = webdriver.Chrome('C:/chromedriver/chromedriver.exe')
        self.num_3=0    #num_3
        self.num_2=28
        self.num_1=8
        self.excel=Excel()
        self.three_division=['']
        self.two_division=['']
        self.one_division=['']
        
    def get_item_name(self,item_name):
    
        self.item_name=item_name
    def find_item_source_first_step_into_page(self):
        
        self.driver.get('https://itemscout.io/')
            
    def find_item_source_second_step_category_1(self):
        '''
        if self.num_2== len(self.two_division):
            self.num_3=0
            self.num_2=0
            self.num_1+=1
        '''
        sleep(3)
        delay=1000
        WebDriverWait(self.driver, delay).until(expected_conditions.presence_of_element_located((By.XPATH, '//*[@id="container"]/table[1]/tr/td[2]/div/div[1]/span')))
        x=self.driver.find_element_by_xpath('//*[@id="container"]/table[1]/tr/td[2]/div/div[1]/span')
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).perform()
        sleep(2)
        delay=1000
        WebDriverWait(self.driver, delay).until(expected_conditions.presence_of_element_located((By.CLASS_NAME, 'itemscout-dropdown-item')))
        self.one_division=self.driver.find_elements_by_class_name('itemscout-dropdown-item')
        
        
        self.one_division[self.num_1].click()
        
        sleep(5)
    def find_item_source_second_step_category_2(self):
        '''
        if self.num_3== len(self.three_division):
            self.num_3=0
            self.num_2+=1
        '''
        delay=1000
        WebDriverWait(self.driver, delay).until(expected_conditions.presence_of_element_located((By.XPATH, '//*[@id="container"]/table[1]/tr/td[2]/div/div[2]/span')))
        x=self.driver.find_element_by_xpath('//*[@id="container"]/table[1]/tr/td[2]/div/div[2]/span')
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).perform()
        
        sleep(3)
        delay=1000
        WebDriverWait(self.driver, delay).until(expected_conditions.presence_of_element_located((By.CLASS_NAME, 'itemscout-dropdown-item')))
        self.two_division=self.driver.find_elements_by_class_name('itemscout-dropdown-item')
        
        self.two_division[self.num_2].click()
        
        sleep(5)
    def find_item_source_second_step_category_3(self):
        delay=1000
        WebDriverWait(self.driver, delay).until(expected_conditions.presence_of_element_located((By.XPATH, '//*[@id="container"]/table[1]/tr/td[2]/div/div[3]/span')))
        x=self.driver.find_element_by_xpath('//*[@id="container"]/table[1]/tr/td[2]/div/div[3]/span')
        webdriver.ActionChains(self.driver).move_to_element(x).perform()
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).perform()
        sleep(3)
        delay=1000
        WebDriverWait(self.driver, delay).until(expected_conditions.presence_of_element_located((By.CLASS_NAME, 'itemscout-dropdown-item')))
        self.three_division=self.driver.find_elements_by_class_name('itemscout-dropdown-item')
        
            
                   
        self.three_division[self.num_3].click()
        self.num_3 +=1
            
        sleep(5)
    def find_item_source_third_step_get_data(self):
        self.list_comp=[]
        self.list_name=[]
        self.list_number=[]
        
        for _ in range(80):
        
            self.driver.execute_script("window.scrollBy(0,{})".format(_))
        
        self.name_tables=self.driver.find_elements_by_xpath('//*[@id="container"]/div[2]/table/tbody/tr/td[2]/a/label')
        self.comp_tables=self.driver.find_elements_by_xpath('//*[@id="container"]/div[2]/table/tbody/tr/td[7]/a/label')
        self.number_tables=self.driver.find_elements_by_xpath('//*[@id="container"]/div[2]/table/tbody/tr/td[6]/a/label')
        for num in range(len(self.name_tables)):
            
            y=self.name_tables[num]
            y=y.get_attribute('innerHTML').strip()
            
            x=self.comp_tables[num]
            x=x.get_attribute('innerHTML').strip()
            
            z=self.number_tables[num]
            z=z.get_attribute('innerHTML').strip()
            h=z.replace(',','')
            if h=='-':
                h=0
            g=x.replace(',','')
            if g=='-':
                g=0
                
            if float(g)<=0.5 and float(h) <=50 :
                self.list_comp.append(x)
                self.list_name.append(y)
                self.list_number.append(z)
                print(y)
                print(x)
                print(z)
    def find_item_source_fourth_step_edit_excel(self,path,sheet):
        self.excel=Excel()
        self.excel.get_execel(path)
        self.excel.get_sheet(sheet)
        for i in range(len(self.list_comp)):
            
            self.excel.edit_excel_data(i+1,1,self.list_name[i])
            self.excel.edit_excel_data(i+1,2,self.list_comp[i])
            self.excel.edit_excel_data(i+1,3,self.list_number[i])
        self.excel.save_excel_data()
        
        
            
        
        
    def get_category(self):
        self.driver.get('https://itemscout.io/keyword/')
        
        sleep(1)
        str_1 = self.item_name
        self.driver.find_element_by_xpath('//*[@id="container"]/div/div[2]/div/input').send_keys(str_1)
        self.driver.find_element_by_xpath('//*[@id="container"]/div/div[2]/div/span').click()
    
        sleep(3)
        try:
            self.driver.find_element_by_xpath('//*[@id="container"]/table/tbody/tr/td[2]/a/label').click()
            delay=100
        except:
            pass
        try :
            WebDriverWait(self.driver, delay).until(expected_conditions.presence_of_element_located((By.XPATH, '//*[@id="container"]/div/div[1]/div[3]/div[2]/div/div[1]')))        
       
            
        
        
            x=self.driver.find_element_by_xpath('//*[@id="container"]/div/div[3]/div[2]/div[1]/div[1]/div[2]/div[1]/div[2]/div/div[1]/a/span[1]/span')
            x=x.get_attribute('innerHTML').strip()
            y=self.driver.find_element_by_xpath('//*[@id="container"]/div/div[3]/div[2]/div[1]/div[1]/div[2]/div[1]/div[2]/div/div[1]/a/span[2]/span[2]')
            y=y.get_attribute('innerHTML').strip()
            z=self.driver.find_element_by_xpath('//*[@id="container"]/div/div[3]/div[2]/div[1]/div[1]/div[2]/div[1]/div[2]/div/div[1]/a/span[3]/b')
            z=z.get_attribute('innerHTML').strip()
            
            print(x+">"+y+">"+z)
            self.category=x+">"+y+">"+z
            print(self.category)
        except:
            print(self.item_name)
            self.delay_case()
            
    def return_category(self):
        return self.category
   
    def delay_case(self):
        self.load_wb=load_workbook(self.excel_path,data_only=True)
           
        self.load_ws = self.load_wb[self.exel_sheet]
        self.load_ws.cell(self.x,self.y).value='원인모름 딜레이'
        self.load_wb.save(self.excel_path)
        
        
    def get_excel_position(self,excel_path,sheet,x,y):
        self.x=x
        self.y=y
        self.excel_path=excel_path
        self.exel_sheet=sheet 
    





class Naver:
    

    def __init__(self):
        self.driver = webdriver.Chrome('C:/chromedriver/chromedriver.exe')
        self.driver.get('https://nid.naver.com/nidlogin.login')
        self.category=''
        self.login()
        #self.save_cookie(driver,path)
        
   

        

        
    def login(self):
      
        ############################################
        id = 
        pw = 
        ############################################
        sleep(0.5)#트래킹 공격이라 인식할수있게때문에 정보에 딜레이를 준다
        self.driver.execute_script("document.getElementsByName('id')[0].value=\'" + id + "\'")
        sleep(1)
        self.driver.execute_script("document.getElementsByName('pw')[0].value=\'" + pw + "\'")
        sleep(1)
        self.driver.find_element_by_xpath('//*[@id="frmNIDLogin"]/fieldset/input').click()
        sleep(2)
         #   def imgcrawling():
         #imgscrawl
        self.window_before = self.driver.window_handles[0]
    def smartstore(self):
        self.driver.switch_to.window(self.window_before) 
        self.driver.get('https://sell.smartstore.naver.com/#/products/create')
        sleep(5)
    def edit_category(self):
        print(self.category)
        
        delay=1000
        WebDriverWait(self.driver, delay).until(expected_conditions.presence_of_element_located((By.XPATH, '//*[@id="productForm"]/ng-include/ui-view[2]/div/div[2]/div/div/div/category-search/div[2]/div/div/div[1]')))
        x=self.driver.find_element_by_xpath('//*[@id="productForm"]/ng-include/ui-view[2]/div/div[2]/div/div/div/category-search/div[2]/div/div/div[1]')
        #driver.execute_script("arguments[0].click();", element)
        
        actions = ActionChains(self.driver).click(x).send_keys(self.category).pause(4).send_keys(Keys.ENTER)
        actions.perform()
        sleep(3)
        delay=1000
        WebDriverWait(self.driver, delay).until(expected_conditions.presence_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/div/button')))
  
        x=self.driver.find_element_by_xpath('/html/body/div[1]/div/div/div[3]/div/button')
        x.click()
    def get_model_name(self,model_name):
        self.model_name=model_name
        #driver.find_element_by_xpath('//*[@id="productForm"]/ng-include/ui-view[2]/div/div[2]/div/div/div/category-search/div[2]/div/div/div[1]').send_keys(category)
    def get_brand_name(self,brand_name):
        self.brand_name=brand_name    
    def get_category(self,category):
        self.category=category
    def get_item_name(self,item_name):
        self.item_name=item_name
    def edit_item_name(self):
        name=self.item_name
        x=self.driver.find_element_by_xpath('//*[@id="productForm"]/ng-include/ui-view[5]/div/div[2]/div/div/div/div/div/div/input')
        x.send_keys(name)
    def get_price(self,price):
        self.price= price
    def edit_price(self):
        x=self.driver.find_element_by_xpath('//*[@id="prd_price2"]')
        x.send_keys(self.price)
    def get_stock(self,stock):
        self.stock=stock
    def edit_stock(self):
        x=self.driver.find_element_by_xpath('//*[@id="stock"]')
        x.send_keys(self.stock)
    def get_tumb_img(self,tumb_img):
        self.tumb_img=tumb_img
    def edit_tumbnail(self):
        
        x=self.driver.find_element_by_xpath('//*[@id="representImage"]/div/div[1]/div/ul/li/div/a/i')
        x.click()
        sleep(3)
        x=self.driver.find_element_by_xpath('/html/body/div[1]/div/div/div/div[2]/div/button[1]')
        x.click()
        sleep(2)
        pyautogui.click(x=1533, y=1057)
        while True:
            try:
                pyautogui.FAILSAFE = False
                print(pyautogui.position())
                print(pyautogui.size())
                center = pyautogui.locateCenterOnScreen('/Users/Administrator/Desktop/ugot/a/address.jpg', grayscale=True)
                self.center1= center 
                pyautogui.click(center)
                
                pyperclip.copy(self.address)
                pyautogui.hotkey("ctrl", "v")
                
                pyautogui.press('enter')
                break
            except:
                pass
        while True:
            try:
                
                center = pyautogui.locateCenterOnScreen('/Users/Administrator/Desktop/ugot/a/filename.jpg', grayscale=True)
                self.center2 = center
                pyautogui.click(center)
                
                
                
                pyperclip.copy(self.tumb_img)
                pyautogui.hotkey("ctrl", "v")
                center = pyautogui.locateCenterOnScreen('/Users/Administrator/Desktop/ugot/a/open.jpg', grayscale=True)
                self.center3 = center
                pyautogui.click(center)
                sleep(3)
                #center = pyautogui.locateCenterOnScreen('/Users/Administrator/Desktop/ugot/a/button.jpg', grayscale=True)
                #self.center4 = center
                #pyautogui.click(center)
                #sleep(2)
                
                break
            except:
                pass
    def get_address_from_img(self,address):
        self.address = address
   
    def go_detail(self):
        x=self.driver.find_element_by_xpath('//*[@id="productForm"]/ng-include/ui-view[10]/div/div[2]/div/div/ncp-se3-form/div[1]/div/p[4]')
        webdriver.ActionChains(self.driver).move_to_element(x).perform()
        for _ in range(20):
            # 가로 = 0, 세로 = 10000 픽셀 스크롤한다.
            self.driver.execute_script("window.scrollBy(0,{})".format(_))
        sleep(1)
        
        
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).perform()
        sleep(3)
     
                               
        self.window_after = self.driver.window_handles[1]
        self.driver.switch_to.window(self.window_after)
        #self.driver.file_detector_context
        #self.driver.execute_cdp_cmd
        #self.driver.switch_to_active_element
        #self.driver.window_handles
        try:
            delay=3
            WebDriverWait(self.driver, delay).until(expected_conditions.presence_of_element_located((By.XPATH,'/html/body/div[5]/div/div/div[2]/a')))
            x=self.driver.find_element_by_xpath('/html/body/div[5]/div/div/div[2]/a')
            webdriver.ActionChains(self.driver).move_to_element(x).click(x).perform()
            sleep(3)

        except:
            pass
        sleep(3)

    def add_img(self,num):
        delay=5
        WebDriverWait(self.driver, delay).until(expected_conditions.presence_of_element_located((By.XPATH,'//*[@id="se_side_comp_list"]/li[2]/button/span')))    
        x=self.driver.find_element_by_xpath('//*[@id="se_side_comp_list"]/li[2]/button/span')
        sleep(1)
        center = pyautogui.locateCenterOnScreen('/Users/Administrator/Desktop/ugot/a/picture.jpg', grayscale=True)
        self.center4 = center
        pyautogui.click(center)
        sleep(2)
        delay=1000
        WebDriverWait(self.driver, delay).until(expected_conditions.presence_of_element_located((By.XPATH, '/html/body/div[5]/div/div/div[2]/div[1]/div/div[1]/button[1]'))) 
        x=self.driver.find_element_by_xpath('/html/body/div[5]/div/div/div[2]/div[1]/div/div[1]/button[1]')
        x.click()
        sleep(4)
        pyautogui.click(self.center2)
        pyperclip.copy(self.tumb_img[:-5]+str(num)+'.jpg')
        pyautogui.hotkey("ctrl", "v")
        pyautogui.click(self.center3)
        sleep(3)
    def exit_detail(self):    
        x=self.driver.find_element_by_xpath('//*[@id="se_top_publish_btn"]')
        
        
        delay=1000
        WebDriverWait(self.driver, delay).until(expected_conditions.presence_of_element_located((By.XPATH, '//*[@id="se_top_publish_btn"]')))
        
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).perform()
    
        sleep(4)
      
        self.driver.switch_to.window(self.window_before)
    def item_info(self):
        for _ in range(20):
            # 가로 = 0, 세로 = 10000 픽셀 스크롤한다.
            self.driver.execute_script("window.scrollBy(0,{})".format(_))
        x=self.driver.find_element_by_xpath('//*[@id="_prod-attr-section"]/div[1]/div/div')
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).perform()
        sleep(2)
        for _ in range(20):
            # 가로 = 0, 세로 = 10000 픽셀 스크롤한다.
            self.driver.execute_script("window.scrollBy(0,{})".format(_))
        delay=1000
        WebDriverWait(self.driver, delay).until(expected_conditions.presence_of_element_located((By.XPATH, '//*[@id="prd-model"]')))    
        x=self.driver.find_element_by_xpath('//*[@id="prd-model"]')
        webdriver.ActionChains(self.driver).move_to_element(x).perform()
        for _ in range(20):
            # 가로 = 0, 세로 = 10000 픽셀 스크롤한다.
            self.driver.execute_script("window.scrollBy(0,{})".format(_))
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).send_keys(self.model_name).perform()
        sleep(2)
        delay=1000
        WebDriverWait(self.driver, delay).until(expected_conditions.presence_of_element_located((By.XPATH, '//*[@id="_prod-attr-section"]/div[2]/div/ncp-naver-shopping-search-info/div[2]/div/div[1]/div/ncp-brand-manufacturer-input/div/div/div/div/div/div[1]/input')))  
        x=self.driver.find_element_by_xpath('//*[@id="_prod-attr-section"]/div[2]/div/ncp-naver-shopping-search-info/div[2]/div/div[1]/div/ncp-brand-manufacturer-input/div/div/div/div/div/div[1]/input')
        webdriver.ActionChains(self.driver).move_to_element(x).perform()
        for _ in range(20):
            # 가로 = 0, 세로 = 10000 픽셀 스크롤한다.
            self.driver.execute_script("window.scrollBy(0,{})".format(_))
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).send_keys(self.brand_name,Keys.ENTER).perform()
        sleep(2)
        x=self.driver.find_element_by_xpath('//*[@id="_prod-attr-section"]/div[2]/div/ncp-naver-shopping-search-info/div[2]/div/div[2]/div/ncp-brand-manufacturer-input/div/div/div/div/div/div[1]/input')
        
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).send_keys(self.brand_name,Keys.ENTER).perform()
        x=self.driver.find_element_by_xpath('//*[@id="_prod-attr-section"]/div[2]/div/div[2]/div/div/label[2]')
        webdriver.ActionChains(self.driver).move_to_element(x).perform()
        for _ in range(20):
            # 가로 = 0, 세로 = 10000 픽셀 스크롤한다.
            self.driver.execute_script("window.scrollBy(0,{})".format(_))
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).perform()
        sleep(3)
        
        x=self.driver.find_element_by_xpath('//*[@id="_prod-attr-section"]/div[2]/div/ncp-origin-area/div/div/div/div/div/div[1]/div[1]/div/div/div[1]')
        webdriver.ActionChains(self.driver).move_to_element(x).perform()
        for _ in range(30):
            # 가로 = 0, 세로 = 10000 픽셀 스크롤한다.
            self.driver.execute_script("window.scrollBy(0,{})".format(_))
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).perform()
        sleep(1)
        center = pyautogui.locateCenterOnScreen('/Users/Administrator/Desktop/ugot/a/onesource.jpg', grayscale=True)
        pyautogui.click(center)
    def nation_delivery_info(self):
        for _ in range(50):
            # 가로 = 0, 세로 = 10000 픽셀 스크롤한다.
            self.driver.execute_script("window.scrollBy(0,{})".format(_))
        x=self.driver.find_element_by_xpath('//*[@id="productForm"]/ng-include/ui-view[15]/div[1]/div/div/div/span')
        webdriver.ActionChains(self.driver).move_to_element(x).perform()
        for _ in range(20):
            # 가로 = 0, 세로 = 10000 픽셀 스크롤한다.
            self.driver.execute_script("window.scrollBy(0,{})".format(_))
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).perform()
        sleep(2)
        x=self.driver.find_element_by_xpath('//*[@id="basic_price"]')
        self.basic_price='3000'
        sleep(2)
        webdriver.ActionChains(self.driver).move_to_element(x).perform()
        for _ in range(20):
            # 가로 = 0, 세로 = 10000 픽셀 스크롤한다.
            self.driver.execute_script("window.scrollBy(0,{})".format(_))
        x.clear()
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).send_keys(self.basic_price).perform()
        x=self.driver.find_element_by_xpath('//*[@id="productForm"]/ng-include/ui-view[15]/div[1]/div[2]/div/div[7]/div[2]/div[5]/div/div[2]/div[4]/div[3]/div/div/input')
        self.productForm= '6000'
        sleep(2)
        webdriver.ActionChains(self.driver).move_to_element(x).perform()
        for _ in range(45):
            # 가로 = 0, 세로 = 10000 픽셀 스크롤한다.
            self.driver.execute_script("window.scrollBy(0,{})".format(_))
        x.clear()
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).send_keys(self.productForm).perform()
        x=self.driver.find_element_by_xpath('//*[@id="productForm"]/ng-include/ui-view[15]/div[1]/div[2]/div/div[7]/div[4]/div/div[1]/div[2]/button')
        webdriver.ActionChains(self.driver).move_to_element(x).perform()
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).perform()
        sleep(4)
        delay=10
        WebDriverWait(self.driver, delay).until(expected_conditions.presence_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[2]/div[2]/div[1]/div/a/p[1]')))
        x=self.driver.find_element_by_xpath('/html/body/div[1]/div/div/div[2]/div[2]/div[1]/div/a/p[1]')
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).perform()        
        sleep(2)
    def nation_return_info(self):
        for _ in range(50):
            # 가로 = 0, 세로 = 10000 픽셀 스크롤한다.
            self.driver.execute_script("window.scrollBy(0,{})".format(_))
        x=self.driver.find_element_by_xpath('//*[@id="productForm"]/ng-include/ui-view[15]/div[2]/div/div/div')
        webdriver.ActionChains(self.driver).move_to_element(x).perform()
        for _ in range(20):
            # 가로 = 0, 세로 = 10000 픽셀 스크롤한다.
            self.driver.execute_script("window.scrollBy(0,{})".format(_))
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).perform()
        sleep(1)        
        x=self.driver.find_element_by_xpath('//*[@id="return_price"]')
        self.return_price='3000'
        webdriver.ActionChains(self.driver).move_to_element(x).perform()
        for _ in range(20):
            # 가로 = 0, 세로 = 10000 픽셀 스크롤한다.
            self.driver.execute_script("window.scrollBy(0,{})".format(_))
        x.clear()    
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).send_keys(self.return_price).perform()
        sleep(1)
        x=self.driver.find_element_by_xpath('//*[@id="exchange_price"]')
        self.exchange_price='6000'
        sleep(1)
        webdriver.ActionChains(self.driver).move_to_element(x).perform()
        for _ in range(20):
            # 가로 = 0, 세로 = 10000 픽셀 스크롤한다.
            self.driver.execute_script("window.scrollBy(0,{})".format(_))
        x.clear()
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).send_keys(self.exchange_price).perform()
        x=self.driver.find_element_by_xpath('//*[@id="productForm"]/ng-include/ui-view[15]/div[2]/div[2]/div/ng-include/div/div[4]/div/div/div[2]/button')
        webdriver.ActionChains(self.driver).move_to_element(x).perform()
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).perform()
        sleep(4)
        delay=10
        WebDriverWait(self.driver, delay).until(expected_conditions.presence_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[2]/div[2]/div[1]/div/a/p[1]'))) 
        x=self.driver.find_element_by_xpath('/html/body/div[1]/div/div/div[2]/div[2]/div[1]/div/a/p[1]')
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).perform()
        sleep(1)
    def view_info(self):
        for _ in range(20):
            # 가로 = 0, 세로 = 10000 픽셀 스크롤한다.
            self.driver.execute_script("window.scrollBy(0,{})".format(_))
        sleep(2)
        x=self.driver.find_element_by_xpath('//*[@id="productForm"]/ng-include/ui-view[19]/div/div/div/div')
        webdriver.ActionChains(self.driver).move_to_element(x).perform()
        
        for _ in range(20):
            # 가로 = 0, 세로 = 10000 픽셀 스크롤한다.
            self.driver.execute_script("window.scrollBy(0,{})".format(_))
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).perform()
        
        
        for _ in range(30):
            # 가로 = 0, 세로 = 10000 픽셀 스크롤한다.
            self.driver.execute_script("window.scrollBy(0,{})".format(_))
        sleep(2)
        while True:
            try:
                x=self.driver.find_element_by_xpath('//*[@id="productForm"]/ng-include/ui-view[19]/div/div[2]/div/div[2]/div/div[2]/ng-include/div/div[3]/div/div/div/label[2]')
                break
            except:
                x=self.driver.find_element_by_xpath('//*[@id="productForm"]/ng-include/ui-view[19]/div/div/div/div')
                webdriver.ActionChains(self.driver).move_to_element(x).click(x).perform()
        webdriver.ActionChains(self.driver).move_to_element(x).perform()
        for _ in range(20):
            # 가로 = 0, 세로 = 10000 픽셀 스크롤한다.
            self.driver.execute_script("window.scrollBy(0,{})".format(_))
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).perform()
    def Tag_setting(self):
        for _ in range(33):
            # 가로 = 0, 세로 = 10000 픽셀 스크롤한다.
            self.driver.execute_script("window.scrollBy(0,{})".format(_))
        while True:
            try:
                x=self.driver.find_element_by_xpath('//*[@id="productForm"]/ng-include/ui-view[18]/div/div[2]/div/div[1]/div/div[3]/div/div/label/span')
                break
            except:
                x=self.driver.find_element_by_xpath('//*[@id="productForm"]/ng-include/ui-view[18]/div/div/div/div')
                webdriver.ActionChains(self.driver).move_to_element(x).click(x).perform()
                sleep(2)
        webdriver.ActionChains(self.driver).move_to_element(x).perform()
        for _ in range(20):
            # 가로 = 0, 세로 = 10000 픽셀 스크롤한다.
            self.driver.execute_script("window.scrollBy(0,{})".format(_))
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).perform()
        sleep(2)
        x=self.driver.find_element_by_xpath('//*[@id="productForm"]/ng-include/ui-view[18]/div/div[2]/div/div[1]/div/div[3]/div[2]/div/div/div[1]')
        self.tag = "프리미엄,고음질,음질좋은,"
        
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).send_keys(self.tag).perform()
        sleep(1)
        
    def save(self):
        for _ in range(20):
            # 가로 = 0, 세로 = 10000 픽셀 스크롤한다.
            self.driver.execute_script("window.scrollBy(0,{})".format(_))
        x=self.driver.find_element_by_xpath('//*[@id="seller-content"]/ui-view/div[3]/div[2]/div[1]/button[3]')
        sleep(2)
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).perform()
        sleep(2)
        x=self.driver.find_element_by_xpath('/html/body/div[1]/div/div/div[2]/div/button[1]')
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).perform()
        sleep(3)
        self.window_after = self.driver.window_handles[1]
        self.driver.switch_to.window(self.window_after)
        self.driver.close()
        sleep(3)        
    def AS_info(self):
        self.as_text='상세페이지 참조'
        delay=1000
        WebDriverWait(self.driver, delay).until(expected_conditions.presence_of_element_located((By.XPATH, '//*[@id="productForm"]/ng-include/ui-view[16]/div/div/div/div')))    
        x=self.driver.find_element_by_xpath('//*[@id="productForm"]/ng-include/ui-view[16]/div/div/div/div')
        webdriver.ActionChains(self.driver).move_to_element(x).perform()
        
        for _ in range(20):
            # 가로 = 0, 세로 = 10000 픽셀 스크롤한다.
            self.driver.execute_script("window.scrollBy(0,{})".format(_))
        webdriver.ActionChains(self.driver).move_to_element(x).click(x).perform()
        delay=1000
        WebDriverWait(self.driver, delay).until(expected_conditions.presence_of_element_located((By.XPATH, '//*[@id="as_info"]'))) 
        x=self.driver.find_element_by_xpath('//*[@id="as_info"]')
        x.clear()
        sleep(1)
        webdriver.ActionChains(self.driver).move_to_element(x).click().send_keys(self.as_text).perform()
        sleep(1)
        
if __name__ == "__main__":
    
    
    
    
    
 
    item_name="Tia Fourte"
    price="100"
    stock='99'
    address="C:/Users/dudnj/yyw/image/"
    tumb_img="[64AUDIO]Tia Fourte_1.jpg"
    brand_name='64AUDIO'
   
    smart = Naver()
    smart.smartstore()
    
    smart.get_category("디지털/가전>음향가전>이어폰")
    smart.edit_category()
    
    smart.get_brand_name(brand_name)
    smart.get_item_name(item_name)
    smart.edit_item_name()
    
    smart.get_price(price)
    smart.edit_price()
    
    smart.get_stock(stock)
    smart.edit_stock()
    
    smart.get_address_from_img(address)
    smart.get_tumb_img(tumb_img)
    
    smart.edit_tumbnail()
    
    smart.go_detail()
    
    for num in range(2,3):
        
        smart.add_img(num)
    
    smart.exit_detail()
    
    smart.item_info()
    smart.nation_delivery_info()
    smart.nation_return_info()
    smart.view_info()
    smart.save()
    
 #   nike = Naver()
