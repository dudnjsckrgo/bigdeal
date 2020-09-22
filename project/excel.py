# -*- coding: utf-8 -*-
"""
Created on Thu Aug 29 12:02:05 2019

@author: Administrator
"""

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import json
import os
import urllib.request
from time import sleep
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from PIL import Image


class Excel:
    
    #def __init__(self):
    def get_execel(self,path):
        #data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
        self.exel_path=path
        self.load_wb = load_workbook(self.exel_path , data_only=True) 
    def get_sheet(self,sheet):
        #시트 이름으로 불러오기
        self.exel_sheet=sheet    
        self.load_ws = self.load_wb[self.exel_sheet]
    def get_A_excel_data(self,y):
        #셀 주소로 값 출력
        self.str_a = self.load_ws['A{}'.format(y)].value
        
    def get_B_excel_data(self,y):
        self.str_b = self.load_ws['B{}'.format(y)].value
    def get_C_excel_data(self,y):
        self.str_c = self.load_ws['C{}'.format(y)].value
    def get_D_excel_data(self,y):
        self.str_d = self.load_ws['D{}'.format(y)].value
    def get_E_excel_data(self,y):
        self.str_e = self.load_ws['E{}'.format(y)].value
    def get_J_excel_data(self,y):
        self.str_j = self.load_ws['J{}'.format(y)].value
    def get_M_excel_data(self,y):
        self.str_m = self.load_ws['M{}'.format(y)].value
    def get_F_excel_data(self,y):
        self.str_f = self.load_ws['F{}'.format(y)].value    
    def get_G_excel_data(self,y):
        self.str_g = self.load_ws['G{}'.format(y)].value
    def get_L_excel_data(self,y):
        self.str_l = self.load_ws['L{}'.format(y)].value
    def get_I_excel_data(self,y):
        self.str_i = self.load_ws['I{}'.format(y)].value    
    def edit_excel_data(self,row,colum,data):
        self.load_ws.cell(row,colum).value= data

    def save_excel_data(self):
        self.load_wb.save(self.exel_path)
        '''
    def Workbook_create(self,workbook_path):
        self.wb=WorkBook()
        self.ws=self.wb.active
        self.wb.save(workbook_path)
        '''
class Kukje:
    def __init__(self):
        self.driver = webdriver.Chrome('C:/chromedriver/chromedriver.exe')
        self.header={'User-Agent':'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) HeadlessChrome/76.0.3809.132 Safari/537.36'}
        self.driver.implicitly_wait(10)       
    def get_item_url(self,item_name_data):
        self.str_b=item_name_data
        self.upper_url=' http://www.all4sound.com/product/search_product.asp?product_name='
        self.item_url = self.upper_url+item_name_data.replace(' ','+')
        self.driver.get(self.item_url)
    def go_detail(self,brand_name_data):
        self.str_a = brand_name_data
        self.driver.implicitly_wait(10) 
       

        for x in self.driver.find_elements_by_xpath('/html/body/table[2]/tbody/tr/td[2]/table/tbody/tr[3]/td/table/tbody/tr[2]/td/table/tbody//a'):
            if self.str_a or self.str_b in x.get_attribute('text'):
                x.click()
                break
            

            
    def get_data(self,upper_path):
        
        self.searchterm="["+self.str_a+"]"+self.str_b 
        self.searchterm2 = self.searchterm +'/'
        self.upper_path= upper_path
        self.rpath= self.upper_path + self.searchterm2
        if not os.path.exists(self.rpath):
            os.mkdir(self.rpath)    
        self.driver.implicitly_wait(10) 
        for _ in range(300):
            # 가로 = 0, 세로 = 10000 픽셀 스크롤한다.
            self.driver.execute_script("window.scrollBy(0,200)")
                
        self.counter = 0
        self.succounter = 0
        for x in self.driver.find_elements_by_xpath('//*[@id="zoom_big_img"]'):
            self.counter = self.counter + 1
            print ("Total Count:", self.counter)
            print ("Succsessful Count:", self.succounter)
            print("URL:",x.get_attribute('src'))
            self.img=x.get_attribute('src')
            imgtype ="jpg"
            first_url ='http://image.all4sound.com/images/product_images/'
            try:
                req = urllib.request.Request(self.img, headers={'User-Agent': self.header})
        
                if self.img[-3:] == imgtype and self.img[:len(first_url)] == first_url :
                    self.succounter = self.succounter + 1 
                    urllib.request.urlretrieve(req.get_full_url(),os.path.join(self.rpath , self.searchterm + "_" + str(self.succounter) + "." + imgtype))
                    self.img_path=self.rpath+self.searchterm+'_'+str(self.succounter)+"."+imgtype
                    image = Image.open(self.img_path)
                    resize_image=image.resize((640, 640))
                    resize_image.save(self.img_path) 
           
            except:
                pass
           
        for x in self.driver.find_elements_by_xpath('/html/body/table[2]/tbody/tr/td[2]/table/tbody/tr[2]/td/table/tbody/tr/td[2]/table[3]/tbody//img'):
            self.counter = self.counter + 1
            print ("Total Count:", self.counter)
            print ("Succsessful Count:", self.succounter)
            print("URL:",x.get_attribute('src'))
            self.img_2=x.get_attribute('src')
            imgtype ="jpg"
            first_url ='http://www.all4sound.com/images/description/'
            try:
                req = urllib.request.Request(self.img_2, headers={'User-Agent': self.header})
        
                if self.img_2[-3:] == imgtype and self.img_2[:len(first_url)] == first_url :
                    self.succounter = self.succounter + 1 
                    urllib.request.urlretrieve(req.get_full_url(),os.path.join(self.rpath , self.searchterm + "_" + str(self.succounter) + "." + imgtype))
            except:
                pass    
        print (self.succounter, "succesfully downloaded")  
          
        
    

if __name__ == "__main__":
    '''
    excel = Excel()
    exel_path ='/Users/Administrator/Downloads/온라인판매20190726.xlsx'
    upper_path="C:/Users/Administrator/Desktop/ugot/a/m/"
    kukje= Kukje()
    excel.get_execel(exel_path)
    excel.get_sheet("Sheet1")
    for i in range(51,190):
        excel.get_A_excel_data(i)
        excel.get_B_excel_data(i)
    
        
        kukje.get_item_url(item_name_data=excel.str_b)
        kukje.go_detail(brand_name_data=excel.str_a)
        kukje.get_data(upper_path)
    '''
    excel = Excel()
    excel.Workbook_create('C:/Users/Administrator/Desktop/ugot/a/m/d/test.xlsx')

   
    